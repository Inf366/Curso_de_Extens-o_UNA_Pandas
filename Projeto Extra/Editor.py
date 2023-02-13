import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Color, Alignment
import matplotlib.pyplot as plt

#Variável local
local = 'DataBase.xlsx'

num = list()
con = 1
while con != 101:
    num.append(con)
    con += 1
plan = pd.read_excel(local)
media = plan["Massa"].mean()
plan["Media"] = ''
plan.at[0,'Media'] = media
MassaA = media-1
#Adiciona a coluna para as Massas do soluto A
massas = list()
for i in plan["Massa"]:
    massas.append(i)
plan["Massa A"] = 0
count = 0
for i in massas:
    M1 = i - 1
    plan.at[count,'Massa A'] = M1
    count += 1
count = 0


#Acha a concentração do soluto A em 1000 ml (1 Litro) e adiciona
#na tabela Concentração Inicial
plan["Concentração Inicial (g/ml)"] = ''
plan.at[0,'Concentração Inicial (g/ml)'] = MassaA


#Cria a concentração com a variação da fração de soluto A
#Ca = Ca0(1-xa)
#Ca = Concentração em Xa; Ca0 = Concentração inicial; xa = Fração de soluto A
plan["Ca"] = 0
x = list()
Ca = list()
ca0 = plan.iloc[0]["Concentração Inicial (g/ml)"]
for i in plan["Xa"]:
    x.append(i)
for i in x:
    plan.at[count,"Ca"] = ca0*(1-i)
    M3 = ca0*(1-i)
    Ca.append(M3)
    count += 1
count = 0


#Calcula a velocidade de reação do soluto A com um k de 0,045 g/(ml*min)
#Velocidade da reação => rA = k[Ca]^alfa , sendo alfa neste caso = 1
A = plan.iloc[0][8]
Alfa = plan.iloc[0][9]
plan["Velocidade"] = 0
ca = list()
for i in plan["Ca"]:
    ca.append(i)
vel = list()
for i in ca:
    M4 = (i**(Alfa))*A
    plan.at[count,"Velocidade"] = (plan.iloc[0]["k"]) * M4
    count += 1
    M2 = (plan.iloc[0]["k"]) * M4
    vel.append(M2)
count = 0
# Drop N/A
plan.dropna(axis='columns',inplace=True)

print(plan)
plan.to_excel("DataBase_Atualizada.xlsx",sheet_name="Solvente A")


#openpyxl

wb = load_workbook("DataBase_Atualizada.xlsx")
ws1 = wb["Solvente A"]
rowmax1 = ws1.max_row
linhas = 2
while linhas <= rowmax1:
    coordenada = 'G'+f'{linhas}'
    if ws1[coordenada].value <= (ca0/2):
        ws1[coordenada].font = Font(color='4169E1')
    linhas += 1
linhas = 2
while linhas <= rowmax1:
    coordenada = 'H'+f'{linhas}'
    if ws1[coordenada].value <= (ca0/2)*0.045:
        ws1[coordenada].font = Font(color='9ACD32')
    linhas += 1

wb.save("DataBase_Atualizada.xlsx")


#Gráficos

#Variação das Massas
plt.plot(num,massas)
plt.xlabel("Valores")
plt.ylabel("Massas (kg)")
plt.title("Variação das Aferições de Massas")
plt.savefig('Variação das Aferições de Massas.jpeg')
plt.close()

#Velocidade x Xa
plt.plot(vel,x)
plt.xlabel("Xa")
plt.ylabel("Velocidade g/(ml*min)")
plt.title("Velocidade vs Xa")
plt.savefig('Velocidade_vs_Xa.jpeg')
plt.close()

#Concentração x Velocidade
plt.plot(vel,num,Ca,num)
plt.xlabel("valores")
plt.ylabel("Velocidade g/(ml*min); Concentração g/ml")
plt.title("Velocidade vs Concentração")
plt.savefig('Velocidade_vs_Concentração.jpeg')
plt.close()
