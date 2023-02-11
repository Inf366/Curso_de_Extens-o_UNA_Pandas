# Autor: GSRS

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sb
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Color, Alignment


Combus = pd.read_excel('ca-2021-02.xlsx')
#plt.hist(Combus['Valor de Venda'])
#plt.title("Preço dos Combustíveis - Nov/2021")
#Rotulo Horizontal
#plt.xlabel("Preço (em reais)")
#plt.ylabel("Quantidades de Coletas")

#Traça uma linha vermelha tracejada com o preço médio
#plt.axvline(Combus['Valor de Venda'].mean(),color='red',linestyle='dashed',linewidth=4)

Combus_mean = Combus['Valor de Venda'].groupby(by=Combus['Produto']).mean()

#Definir a área da figura
#plt.figure(figsize=(10,5))
#Combus_mean.plot(kind="barh",
#                 xlabel="Tipo de Combustível",
#                 ylabel="Preço em reais/litro",
#                 title="Média de Preços por Combustíveis",
#                 alpha=0.3,
#                 color="red")
#plt.grid()


#Pós intervalo

#remover as linhas do gráfico
#sb.despine()
#plt.show()
#Saida = "Por_Litro.xlsx"
#Combus_mean.to_excel(Saida,"sumário")
#print("Feito")

#Abre o excel no openpyxl
wb = load_workbook("Por_Litro.xlsx")
#Pegar a planilha certa usando Sheet Name
ws = wb['sumário']
#Pintar o cabeçalho da tabela de "Cinzinha"
Cinza = PatternFill("solid",fgColor="cccccc")
coords = ['A1','B1']
for coord in coords:
    ws[coord].fill = Cinza

#Salvar o Excel

#Onde o preço do combustível for maior ou igual a 6,5 reais, pinta a fonte
#de vermelho e deixa negrito
MAXROW = ws.max_row
num_linha = 2
while num_linha <= MAXROW:
    coordenada = 'B'+f'{num_linha}'
    if ws[coordenada].value >= 6.5:
        ws[coordenada].font = Font(bold=True, color='FF0000')
    num_linha += 1
wb.save("Por_Litro.xlsx")
print("Feito")
